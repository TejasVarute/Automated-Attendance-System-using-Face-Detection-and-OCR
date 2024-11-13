import os
import cv2
import time
import threading
import numpy as np
import customtkinter
import openpyxl as op
from Attendance_Process import *
from tkinter import filedialog
from tkinter import messagebox

customtkinter.set_appearance_mode("System")  #Modes: "System" (standard), "Dark", "Light"
customtkinter.set_default_color_theme("dark-blue")  #Themes: "blue" (standard), "green", "dark-blue"

#class Password
class Password():
    def __init__(self, folder_path, data_file): 
        self.folder_path = folder_path
        self.data_file = data_file
        self.data_book = op.load_workbook(data_file)
        self.data_sheet = self.data_book.active

        def password_incorrect():
            ctk = customtkinter.CTk()
            width_of_window = 250
            height_of_window = 75
            screen_width = ctk.winfo_screenwidth()
            screen_height = ctk.winfo_screenheight()
            x_coordinate = (screen_width/2)-(width_of_window/2)
            y_coordinate = (screen_height/2)-(height_of_window/2)
            ctk.geometry("%dx%d+%d+%d" %(width_of_window,height_of_window,x_coordinate,y_coordinate))
            ctk.overrideredirect(True)

            customtkinter.CTkFrame(ctk, width=250, height=75, fg_color='transparent', corner_radius=5,  border_width=5).pack(padx = 0, pady = 0)
            label = customtkinter.CTkLabel(ctk, text='\nIncorrect Password', font=('Calibri',18,'bold'))
            label.place(x=50, y=5)
            ctk.after(2000, ctk.destroy)
            ctk.mainloop()

        def password_saved():
            def close():
                self.data_book.save(self.data_file)
                ctk.destroy()
                exit(0)

            ctk = customtkinter.CTk()
            width_of_window = 250
            height_of_window = 75
            screen_width = ctk.winfo_screenwidth()
            screen_height = ctk.winfo_screenheight()
            x_coordinate = (screen_width/2)-(width_of_window/2)
            y_coordinate = (screen_height/2)-(height_of_window/2)
            ctk.geometry("%dx%d+%d+%d" %(width_of_window,height_of_window,x_coordinate,y_coordinate))
            ctk.overrideredirect(True)

            customtkinter.CTkFrame(ctk, width=250, height=75, fg_color='transparent', corner_radius=5,  border_width=5).pack(padx = 0, pady = 0)
            label = customtkinter.CTkLabel(ctk, text='\nPassword Saved', font=('Calibri',18,'bold'))
            label.place(x=70, y=5)
            ctk.after(2000, close)
            ctk.mainloop()
            
        def change_password():
            def change():
                old_password = old_entry_password.get()

                if old_password != str(self.data_sheet['B1'].value):
                    password_incorrect()
                    ctk.after(1000, ctk.destroy)
                    exit(0)

                ctk.after(0, label_password.destroy)
                ctk.after(0, old_entry_password.destroy)
                ctk.after(0, submit_button.destroy)

                def set_new_pass():
                    new_pass = new_entry_password.get()
                    self.data_sheet['B1'] = int(new_pass)
                    password_saved()

                new_label_password = customtkinter.CTkLabel(ctk, text='Enter New Password', corner_radius= 15, bg_color='transparent', font=('Calibri',18,'bold'))
                new_label_password.place(x=65, y=15)

                new_entry_password = customtkinter.CTkEntry(ctk, show='*', corner_radius= 15, bg_color='transparent')
                new_entry_password.place(x=85, y=60)

                new_submit_button = customtkinter.CTkButton(ctk, text='Submit', corner_radius= 15, bg_color='transparent', border_width=2, command=set_new_pass)
                new_submit_button.place(x=85, y=95)

            ctk = customtkinter.CTk()
            width_of_window = 300
            height_of_window = 150
            screen_width = ctk.winfo_screenwidth()
            screen_height = ctk.winfo_screenheight()
            x_coordinate = (screen_width/2)-(width_of_window/2)
            y_coordinate = (screen_height/2)-(height_of_window/2)
            ctk.geometry("%dx%d+%d+%d" %(width_of_window,height_of_window,x_coordinate,y_coordinate))
            ctk.overrideredirect(True)
            ctk.columnconfigure((0,1,2), weight=0)
            ctk.rowconfigure((0,1,2), weight=0)

            customtkinter.CTkFrame(ctk, width=300, height=150, fg_color='transparent', corner_radius=5,  border_width=5).pack(padx = 0, pady = 0)

            old_label_password = customtkinter.CTkLabel(ctk, text='Enter old Password', corner_radius= 15, bg_color='transparent', font=('Calibri',18,'bold'))
            old_label_password.place(x=65, y=15)

            old_entry_password = customtkinter.CTkEntry(ctk, show='*', corner_radius= 15, bg_color='transparent')
            old_entry_password.place(x=85, y=60)

            old_submit_button = customtkinter.CTkButton(ctk, text='Submit', corner_radius= 15, bg_color='transparent', border_width=2, command=change)
            old_submit_button.place(x=85, y=95)

            ctk.mainloop()

        def get_password():
            password = entry_password.get()
            if password == str(self.data_sheet['B1'].value):
                ctk.destroy()
            else:
                password_incorrect()

        ctk = customtkinter.CTk()

        width_of_window = 310
        height_of_window = 150
        screen_width = ctk.winfo_screenwidth()
        screen_height = ctk.winfo_screenheight()
        x_coordinate = (screen_width/2)-(width_of_window/2)
        y_coordinate = (screen_height/2)-(height_of_window/2)
        ctk.geometry("%dx%d+%d+%d" %(width_of_window,height_of_window,x_coordinate,y_coordinate))
        ctk.overrideredirect(True)
        ctk.columnconfigure((0,1,2), weight=0)
        ctk.rowconfigure((0,1,2), weight=0)

        customtkinter.CTkFrame(ctk, width=310, height=150, fg_color='transparent', corner_radius=5,  border_width=5).pack(padx = 0, pady = 0)

        label_password = customtkinter.CTkLabel(ctk, text='Enter Password', corner_radius= 15, bg_color='transparent', font=('Calibri',18,'bold'))
        label_password.place(x=85, y=15)

        entry_password = customtkinter.CTkEntry(ctk, show='*', corner_radius= 15, bg_color='transparent')
        entry_password.place(x=85, y=60)

        submit_button = customtkinter.CTkButton(ctk, text='Submit', corner_radius= 15, bg_color='transparent', border_width=2, command=get_password)
        submit_button.place(x=10, y=100)

        change_button = customtkinter.CTkButton(ctk, text='change password', corner_radius= 15, bg_color='transparent', border_width=2, command=change_password)
        change_button.place(x=160, y=100)

        ctk.mainloop()

#class 5 GUI
class App():
    def __init__(self, folder_path, data_file):
        self.folder_path = folder_path
        self.data_file = data_file
        self.image_path = None

        self.data_book = op.load_workbook(self.data_file)
        self.data_sheet = self.data_book.active

        self.max_roll = self.data_sheet['B3'].value
        self.drivers_list = self.select_camera()
        self.rolls = []
        self.rolls.append(str(self.max_roll))

        self.cam_driver = 0
        self.face_count = 0
        self.sticker_count = 0

    def select_camera(self):
        cameras = []
        for i in range(5):
            cap = cv2.VideoCapture(i)
            if cap.isOpened():
                cameras.append(f'Driver : {str(i+1)}')
                cap.release()

        if len(cameras) == 0:
            return ['No camera detected']
        else:
            return cameras

    def end_popup(self): 
        p_self = customtkinter.CTk()
        p_self.title("Beyond Rollcall")
        width_of_window = 750
        height_of_window = 350
        screen_width = p_self.winfo_screenwidth()
        screen_height = p_self.winfo_screenheight()
        x_coordinate = (screen_width/2)-(width_of_window/2)
        y_coordinate = (screen_height/2)-(height_of_window/2)
        p_self.geometry("%dx%d+%d+%d" %(width_of_window,height_of_window,x_coordinate,y_coordinate))
        p_self.resizable(0,0)

        p_self.grid_columnconfigure((0,2), weight=0)
        p_self.grid_columnconfigure((1), weight=1)
        p_self.grid_rowconfigure((0,1,2,3,4,5,6), weight=0)
        #p_self.grid_rowconfigure((5), weight=0)

        p_self.empty1 = customtkinter.CTkLabel(p_self, text = " ")
        p_self.empty1.grid(row=0, column=1, sticky='n', pady=10, padx=10)
        p_self.empty2 = customtkinter.CTkLabel(p_self, text = " ")
        p_self.empty2.grid(row=2, column=0, sticky='n', pady=10, padx=10)
        p_self.empty3 = customtkinter.CTkLabel(p_self, text = " ")
        p_self.empty3.grid(row=2, column=1, sticky='n', pady=10, padx=10)

        p_self.title = customtkinter.CTkLabel(p_self, text = "Attendance Successfully Marked", font=customtkinter.CTkFont(size=20, weight="bold"))
        p_self.title.grid(row=1, column=1, sticky='n', pady=10)

        p_self.label_1 = customtkinter.CTkLabel(p_self, text=f"Face detected   :    {self.face_count}", font=customtkinter.CTkFont(size=16, weight="bold"))
        p_self.label_1.grid(row=3, column=1, sticky = "we", pady=5, padx=80)
        p_self.label_2 = customtkinter.CTkLabel(p_self, text=f"Roll Number detected   :    {self.sticker_count}", font=customtkinter.CTkFont(size=16, weight="bold"))
        p_self.label_2.grid(row=4, column=1, sticky = "we", pady=5, padx=80)

        p_self.label_3 = customtkinter.CTkLabel(p_self, text="\t\tPresent Numbers      : ", font=customtkinter.CTkFont(size=16, weight="bold"))
        p_self.label_3.grid(row=5, column=1, padx=10, pady=15, sticky = "w")
        p_self.open_button = customtkinter.CTkButton(p_self, text="Open", corner_radius= 15 , border_width=2, command=self.open_present_no_file)
        p_self.open_button.grid(row=5, column=1, padx=200, pady=15, sticky="e")

        p_self.exit_botton = customtkinter.CTkButton(p_self, text="Exit", corner_radius= 15, border_width=2, command=p_self.destroy)
        p_self.exit_botton.grid(row=6, column=1, padx=20, pady=20, sticky="s")
        
        p_self.mainloop()
    
    def loading_popup(self):
            def endcall():
                ctk.destroy()
                self.end_popup()

            def detection_call():
                detect = Detection(self.image_path, self.folder_path, self.max_roll, self.data_sheet, self.data_file)
                self.face_count, self.sticker_count = detect.process()

                ctk.after(0, loading.destroy)
                ctk.after(0, progressbar_3.destroy)

                done = customtkinter.CTkLabel(ctk, text="Process Done", font=("Arial", 20))
                done.place(x=200, y=15)
                exit_botton = customtkinter.CTkButton(ctk, text="Exit", font=("Arial", 14) ,corner_radius= 15, border_width=2 ,command=endcall)
                exit_botton.place(x=192, y=60)
                ctk.update_idletasks()

            def start_loading():
                for i in range(101):
                    progressbar_3.set(i/100)
                    ctk.update_idletasks()
                    time.sleep(0.04)

            ctk = customtkinter.CTk()
            width_of_window = 520
            height_of_window = 100
            screen_width = ctk.winfo_screenwidth()
            screen_height = ctk.winfo_screenheight()
            x_coordinate = (screen_width/2)-(width_of_window/2)
            y_coordinate = (screen_height/2)-(height_of_window/2)
            ctk.geometry("%dx%d+%d+%d" %(width_of_window,height_of_window,x_coordinate,y_coordinate))
            ctk.title('Progress Bar')
            ctk.overrideredirect(True)

            customtkinter.CTkFrame(ctk, width=520, height=100, fg_color='transparent', corner_radius=5,  border_width=5).pack(padx = 0, pady = 0)

            loading = customtkinter.CTkLabel(ctk, text="Loading...", font=("Arial", 20))
            loading.place(x=235, y=20)

            progressbar_3 = customtkinter.CTkProgressBar(ctk, width=500, height=12, corner_radius=25, mode="determinate")
            progressbar_3.set(0)
            ctk.update_idletasks()
            progressbar_3.place(x=10, y=70)

            ctk.after(100, start_loading)
            processing_thread = threading.Thread(target=detection_call)
            processing_thread.start()
            ctk.mainloop()

    def capture_img(self):
        messagebox.showinfo(title="Instruction", message="Q : Quit\n R : Retake\n Space : Accept")
        delay = 3
        flag = 0
        cap = cv2.VideoCapture(self.cam_driver)
        cap.set(cv2.CAP_PROP_FRAME_WIDTH, 1920)
        cap.set(cv2.CAP_PROP_FRAME_HEIGHT, 1080)

        if not cap.isOpened():
            return
        start_time = cv2.getTickCount()
        
        while True:
                ret, frame = cap.read()
                current_time = cv2.getTickCount()
                elapsed_time = (current_time - start_time) / cv2.getTickFrequency()
                
                if elapsed_time >= delay:
                    break
                if cv2.waitKey(1) & 0xFF == ord('q'):
                    break
                cv2.imshow('Captured Image', frame)

        key = cv2.waitKey(0)
        if key == ord(' '):
            #Enhacing captured image
            brightness = 5
            contrast = 5
            adjusted_frame = cv2.convertScaleAbs(frame, alpha=1 + contrast / 100.0, beta=brightness)
            sharpening_kernel = np.array([[-1, -1, -1],
                                    [-1,  9, -1],
                                    [-1, -1, -1]])
            enhanced_frame = cv2.filter2D(adjusted_frame, -1, sharpening_kernel)
            
            self.image_path = f'{self.folder_path}\\Captured_image.jpg'
            cv2.imwrite(self.image_path, enhanced_frame)
            flag = 1
        elif key == ord('r'):
            cap.release()
            cv2.destroyAllWindows()
            self.capture_img()
            flag = 1
        elif key == ord('q'):
            flag = 0
            cap.release()
            cv2.destroyAllWindows()
            exit(0)
            
        cap.release()
        cv2.destroyAllWindows()
    
        if flag == 1:
            self.loading_popup()

    def select_path(self):
        self.image_path = filedialog.askopenfilename(initialdir='/', title='Open image file', filetypes=(('Image File', ('*.jpg', '*.png', '*.jpeg')), ('All Files','*.*')))
        if self.image_path != '':
            self.loading_popup()
    
    def file_not_create(self):
        ctk = customtkinter.CTk()
        width_of_window = 250
        height_of_window = 75
        screen_width = ctk.winfo_screenwidth()
        screen_height = ctk.winfo_screenheight()
        x_coordinate = (screen_width/2)-(width_of_window/2)
        y_coordinate = (screen_height/2)-(height_of_window/2)
        ctk.geometry("%dx%d+%d+%d" %(width_of_window,height_of_window,x_coordinate,y_coordinate))
        ctk.overrideredirect(True)

        customtkinter.CTkFrame(ctk, width=250, height=75, fg_color='transparent', corner_radius=5,  border_width=5).pack(padx=0, pady=0)
        label = customtkinter.CTkLabel(ctk, text='\nFile not Created Yet', font=('Calibri',18,'bold'))
        label.place(x=50, y=5)
        ctk.after(2000, ctk.destroy)
        ctk.mainloop()

    def open_database(self):
        try:
            tmp_data_book = op.load_workbook(self.data_file)
            tmp_sheet = tmp_data_book.active
            count = int(tmp_sheet['B2'].value)
            tmp_data_book.save(self.data_file)

            self.database_file = f'{self.folder_path}\\Attendance Data {count}.xlsx'
            os.startfile(self.database_file)
        except Exception:
            self.file_not_create()
        
    def result_file(self):
        try:
            if(bool(open(f'{self.folder_path}\\Result.jpg')) == True):
                self.result_file_path = f'{self.folder_path}\\Result.jpg'
                os.startfile(self.result_file_path)
        except Exception:
            self.file_not_create()
    
    def open_present_no_file(self):
        os.startfile(f'{self.folder_path}/presenty.txt')

    def main(self, loading_window):
        loading_window.destroy()
        Password(self.folder_path, self.data_file)
        
        [self.rolls.append(str(ele)) for ele in range (1, 101)]

        def roll_submit():
            self.max_roll = c_self.combobox_1_1.get()
            self.max_roll = int(self.max_roll)
            self.data_sheet['B3'] = self.max_roll
            self.data_book.save(self.data_file)

        def set_driver():
            tmp = c_self.optionbox_1.get()
            self.cam_driver = (int(tmp[-1])-1)
        
        def close():
            c_self.destroy()
            exit(0)
        
        def change_appearance_mode_event(new_appearance_mode: str):
            customtkinter.set_appearance_mode(new_appearance_mode)

        #defination
        c_self = customtkinter.CTk()
        c_self.title("Beyond Rollcall")
        current = os.path.dirname(__file__)
        c_self.iconbitmap(f"{current}/icon.ico")
        width_of_window = 900
        height_of_window = 550
        screen_width = c_self.winfo_screenwidth()
        screen_height = c_self.winfo_screenheight()
        x_coordinate = (screen_width/2)-(width_of_window/2)
        y_coordinate = (screen_height/2)-(height_of_window/2)
        c_self.geometry("%dx%d+%d+%d" %(width_of_window,height_of_window,x_coordinate,y_coordinate))
        #c_self.resizable(0,0)

        c_self.grid_columnconfigure((0,2), weight=0)
        c_self.grid_columnconfigure((1), weight=1)
        c_self.grid_rowconfigure((0, 1, 2, 3, 4), weight=0)
        
        #c_self.overrideredirect(True)                              #remove edge frame

        #title set
        c_self.empty = customtkinter.CTkLabel(c_self, text = " ")
        c_self.empty.grid(row=0, column=1, sticky='nsew', pady=20)
        c_self.title = customtkinter.CTkLabel(c_self, text = "Beyond the roll call: Smart strategies for Attendance Management", font=customtkinter.CTkFont('Game of Squids', size=20, weight="bold"))
        c_self.title.grid(row=1, column=1, sticky='nsew', pady=20)

        #subtab set
        c_self.tabview = customtkinter.CTkTabview(c_self, width=100, height=100, corner_radius= 15, border_width=5)
        c_self.tabview.grid(row=2, column=1, padx=(20, 0), pady=(20, 0), sticky="n")
        c_self.tabview.add("Setting")
        c_self.tabview.add("Make Attendance")
        c_self.tabview.add("Database")

        #declaration contains in subtab
        c_self.tabview.tab("Setting").grid_columnconfigure((0,2), weight=1)
        c_self.tabview.tab("Setting").grid_columnconfigure((1), weight=0)
        c_self.tabview.tab("Make Attendance").grid_columnconfigure((0,2), weight=1)
        c_self.tabview.tab("Make Attendance").grid_columnconfigure((1), weight=0)
        c_self.tabview.tab("Database").grid_columnconfigure((0,2), weight=1)
        c_self.tabview.tab("Database").grid_columnconfigure((1), weight=0)

        #selection strength set
        c_self.label_1_1 = customtkinter.CTkLabel(c_self.tabview.tab("Setting"), text="Select class strength : ", font=customtkinter.CTkFont(size=14))
        c_self.label_1_1.grid(row=0, column=0, padx=20, pady=20, sticky = "n")
        c_self.combobox_1_1 = customtkinter.CTkComboBox(c_self.tabview.tab("Setting") , corner_radius= 15 ,values=self.rolls, width=150)
        c_self.combobox_1_1.grid(row=0, column=1, padx=20, pady=(10, 10))
        c_self.set_button_1_1 = customtkinter.CTkButton(c_self.tabview.tab("Setting"), text="Set", corner_radius= 15, bg_color='transparent', border_width=2, command=roll_submit)
        c_self.set_button_1_1.grid(row=0, column=2, padx=20, pady=(10, 10), sticky="w")

        #selection camera driver set
        c_self.label_1_2 = customtkinter.CTkLabel(c_self.tabview.tab("Setting"), text="Select camera driver : ", font=customtkinter.CTkFont(size=14))
        c_self.label_1_2.grid(row=1, column=0, padx=20, pady=20, sticky = "n")
        c_self.optionbox_1 = customtkinter.CTkOptionMenu(c_self.tabview.tab("Setting") , corner_radius= 15 ,values=self.drivers_list, width=150)
        c_self.optionbox_1.grid(row=1, column=1, padx=20, pady=(10, 10))
        c_self.set_button_1_2 = customtkinter.CTkButton(c_self.tabview.tab("Setting"), text="Set", corner_radius= 15, bg_color='transparent', border_width=2, command=set_driver)
        c_self.set_button_1_2.grid(row=1, column=2, padx=20, pady=(10, 10), sticky="w")

        #capture set
        c_self.label_2_1 = customtkinter.CTkLabel(c_self.tabview.tab("Make Attendance"), text="Capture Image : ", font=customtkinter.CTkFont(size=14))
        c_self.label_2_1.grid(row=0, column=0, padx=20, pady=20, sticky = "n")
        c_self.cap_button = customtkinter.CTkButton(c_self.tabview.tab("Make Attendance"), text="Capture", corner_radius= 15, bg_color='transparent', border_width=2, command=self.capture_img)
        c_self.cap_button.grid(row=0, column=1, padx=20, pady=(10, 10), sticky="w")

        #appearance mode set
        c_self.appearance_mode_label = customtkinter.CTkLabel(c_self.tabview.tab("Setting"), text="Appearance Mode", font=customtkinter.CTkFont(size=14))
        c_self.appearance_mode_label.grid(row=2, column=0, padx=20, pady=20)
        c_self.appearance_mode_optionemenu = customtkinter.CTkOptionMenu(c_self.tabview.tab("Setting"), corner_radius=15, bg_color='transparent', values=["System","Light", "Dark"],command=change_appearance_mode_event)
        c_self.appearance_mode_optionemenu.grid(row=2, column=1, padx=20, pady=20, sticky="n")

        #select file set
        c_self.label_3_1 = customtkinter.CTkLabel(c_self.tabview.tab("Make Attendance"), text="Select Path : ", font=customtkinter.CTkFont(size=14))
        c_self.label_3_1.grid(row=1, column=0, padx=20, pady=20, sticky = "n")
        c_self.select_button = customtkinter.CTkButton(c_self.tabview.tab("Make Attendance"), text="Select" , corner_radius= 15, bg_color='transparent', border_width=2, command=self.select_path)
        c_self.select_button.grid(row=1, column=1, padx=20, pady=(10, 10), sticky="w")

        #open database file set
        c_self.label_4_1 = customtkinter.CTkLabel(c_self.tabview.tab("Database"), text="Open database File : ", font=customtkinter.CTkFont(size=14))
        c_self.label_4_1.grid(row=1, column=0, padx=20, pady=20, sticky = "n")
        c_self.open_button_4_1 = customtkinter.CTkButton(c_self.tabview.tab("Database"), text="Open", corner_radius= 15, bg_color='transparent', border_width=2, command=self.open_database)
        c_self.open_button_4_1.grid(row=1, column=1, padx=20, pady=(10, 10), sticky="w")

        #open result file set
        c_self.label_4_2 = customtkinter.CTkLabel(c_self.tabview.tab("Database"), text="Open recent result File : ", font=customtkinter.CTkFont(size=14))
        c_self.label_4_2.grid(row=2, column=0, padx=20, pady=20, sticky = "n")
        c_self.open_button_4_2 = customtkinter.CTkButton(c_self.tabview.tab("Database"), text="Open", corner_radius= 15, bg_color='transparent', border_width=2, command=self.result_file)
        c_self.open_button_4_2.grid(row=2, column=1, padx=20, pady=(10, 10), sticky="w")

        #exit button set
        c_self.exit_botton = customtkinter.CTkButton(c_self, text="Exit", corner_radius= 15, bg_color='transparent', border_width=2, command=close)
        c_self.exit_botton.grid(row=3, column=1, padx=20, pady=20, sticky="n")

        c_self.mainloop()

#class 5 PreLoading Screen    
class PreLoading():
    def __init__(self):
        self.ct = customtkinter
        self.folder_path = ''
        self.data_file = ''

        def data_file(folder_path):
            if not os.path.exists(f'{folder_path}\\Do not delete.xlsx'):
                wb = op.Workbook()
                sheet = wb.active

                sheet['A1'] = 'Password'
                sheet['A2'] = 'Attendance files count'
                sheet['A3'] = 'Max Roll No.'

                sheet['B1'] = 123456
                sheet['B2'] = 1
                sheet['B3'] = 1

                wb.save(f'{folder_path}\\Do not delete.xlsx')
            self.data_file =  f'{folder_path}\\Do not delete.xlsx'

        def check_folder():
            try:
                current = os.path.dirname(__file__)
                os.makedirs(os.path.expanduser(f'{current}/Attendance System'), exist_ok=True)
                self.folder_path = os.path.expanduser(f'{current}/Attendance System')
            except Exception:
                current = os.path.dirname(__file__)
                new_folder_path = os.path.join(current, "Attendance System")
                os.makedirs(new_folder_path, exist_ok=True)
                self.folder_path = new_folder_path

        check_folder()
        data_file(self.folder_path)

    def loading(self):
        def process():
            a = App(self.folder_path, self.data_file)
            a.main(ctk)
        
        def start_anim():
            p_bar1.place(x=120, y=210)
            p_bar1.start()

        ct = self.ct
        ctk = ct.CTk()
        width_of_window = 426
        height_of_window = 250
        screen_width = ctk.winfo_screenwidth()
        screen_height = ctk.winfo_screenheight()
        x_coordinate = (screen_width/2)-(width_of_window/2)
        y_coordinate = (screen_height/2)-(height_of_window/2)
        ctk.geometry("%dx%d+%d+%d" %(width_of_window,height_of_window,x_coordinate,y_coordinate))
        ctk.overrideredirect(True)

        customtkinter.CTkFrame(ctk, width=426, height=250, fg_color='transparent', corner_radius=5,  border_width=5).pack(padx = 0, pady = 0)
        label1 = ct.CTkLabel(ctk, text='Classroom Attendance System', font=('Game of Squids', 24, 'bold'))
        label1.place(x=40, y=90)

        label2 = ct.CTkLabel(ctk, text='Loading...', font=('Calibri', 16))
        label2.place(x=190, y=175)

        p_bar1 = ct.CTkProgressBar(ctk, width=200, height=9, progress_color='gray' ,fg_color='white', corner_radius=30, mode='indeterminate', border_width=1)

        ctk.after(10, start_anim)
        ctk.after(7000, process)
        ctk.mainloop()

p = PreLoading()
p.loading()

import os
import cv2
import time
import easyocr
import openpyxl as op
from ultralytics import YOLO

#class 1
class DBMS():
    def __init__(self, roll_lst, max_roll, folder_path, data_sheet, data_file):
        self.folder_path = folder_path
        self.roll_lst = roll_lst

        self.data_file = data_file
        self.data_sheet = data_sheet

        self.max_roll = max_roll
        self.today = time.strftime('%a')
        self.db_path = None
        self.count = None
        self.db_location()

    def file_update(self):
        self.count = int(self.data_sheet['B2'].value)

        tmp_data_book = op.load_workbook(self.data_file)
        tmp_sheet = tmp_data_book.active
        update = int(self.count)
        tmp_sheet['B2'] = update + 1
        tmp_data_book.save(self.data_file)

    def db_location(self):
        tmp_data_book = op.load_workbook(self.data_file)
        tmp_sheet = tmp_data_book.active
        self.count = tmp_sheet['B2'].value
        tmp_data_book.save(self.data_file)

        self.db_path = f'{self.folder_path}\\Attendance Data {self.count}.xlsx'

    def create_db(self):
        wb = op.Workbook()
        wb.save(self.db_path)

    def db_check(self):
        if not os.path.exists(self.db_path):
            self.create_db()

    def day_result(self, sheet):
        sheet.cell(row=sheet.min_row, column=sheet.max_column+1).value = 'Result'
        for row in range (sheet.min_row+1, sheet.max_row + 1):
            present_count=0
            total_count=0
            for col in range(1, sheet.max_column + 1):
                cell_value = sheet.cell(row=row, column=col).value
                if cell_value == 'Present':
                    present_count += 1
                if cell_value == 'Present' or cell_value == 'Absent':
                    total_count += 1
            if total_count > 0:
                percentage = (present_count / total_count) * 100
            else:
                percentage = 0
            sheet.cell(row=row, column=sheet.max_column, value=percentage)

    def week_result(self, wb):
        self.day_result(sheet=wb[wb.sheetnames[len(wb.sheetnames)-1]])
        result_sheet = wb.create_sheet(title = 'Week_Report')
        result_sheet = wb['Week_Report']

        result_sheet.cell(row=result_sheet.min_row, column=result_sheet.min_column, value='Result')
        tmp_sheet = wb.active
        for row in range (tmp_sheet.min_row+1, tmp_sheet.max_row + 1):
            total=0
            for ele in wb.sheetnames:
                sheet = wb[ele]
                if sheet.cell(row=sheet.min_row, column=sheet.max_column).value == 'Result':
                    cell_value = sheet.cell(row=row, column=sheet.max_column).value
                    if cell_value != None and cell_value != 0:
                        total += cell_value
            total = total / (len(wb.sheetnames)-1)
            result_sheet.cell(row=row, column=result_sheet.min_column, value = total)
        wb.save(self.db_path)
        self.file_update()

    def check_day(self, sheet, wb):
        if sheet.title == 'Sheet':
            sheet.title = self.today
        elif self.today in wb.sheetnames:
            sheet = wb[self.today]
        elif self.today not in wb.sheetnames:
            self.day_result(wb[wb.sheetnames[len(wb.sheetnames)-1]])
            sheet = wb.create_sheet(title = self.today)
        return sheet

    def database(self):
        self.db_check()
        wb=op.load_workbook(self.db_path)
        sheet = wb.active
        if self.today == wb.sheetnames[0] and len(wb.sheetnames) > 1:
            self.week_result(wb)
            self.db_location()
            self.db_check()
            wb=op.load_workbook(self.db_path)
            sheet = wb.active
        
        sheet = self.check_day(sheet, wb)

        if sheet.cell(row=sheet.max_row, column=sheet.min_column).value != self.max_roll:
            sheet.cell(row=sheet.min_row, column=sheet.min_column).value = 'Roll No'
            sheet.cell(row=sheet.min_row, column=sheet.max_column+1).value = str(time.strftime("%x %H:%M"))
            for i in range (1,((self.max_roll)+1)):
                sheet.append([i])
        elif sheet.cell(row=sheet.min_row, column= sheet.max_column).value != str(time.strftime("%x %H:%M")):
            sheet.cell(row=sheet.min_row, column=sheet.max_column+1).value = str(time.strftime("%x %H:%M"))
        wb.save(self.db_path)
    
    def process(self):
        self.database()
        wb=op.load_workbook(self.db_path)
        sheet=wb.active
        sheet = self.check_day(sheet, wb)

        for i in range (2, sheet.max_row+1):
            if sheet.cell(row=i, column=1).value in self.roll_lst :
                sheet.cell(row=i, column=sheet.max_column).value = "Present"
            else:
                sheet.cell(row=i, column=sheet.max_column).value = "Absent"
        wb.save(self.db_path)

#class 2
class Detection():
    def __init__(self, image_path, folder_path, max_roll, data_sheet, data_file):
        self.model_path = os.path.dirname(__file__)
        self.image_path = image_path
        self.folder_path = folder_path
        self.max_roll = max_roll
        self.data_sheet = data_sheet
        self.data_file = data_file
        
        self.reader = easyocr.Reader(['en'], gpu=False)
        self.face_detector = YOLO(f'{self.model_path}\\model\\face_detection.pt')
        
        self.storage = set()
        self.face_count = 0
        
        self.image = cv2.imread(self.image_path)

    def find_and_display_numbers(self, sticker_region):
        sticker_results = self.reader.readtext(image=sticker_region, adjust_contrast=0.5)
        for box_size, text, score in sticker_results:
            if text.isnumeric():
                self.storage.add(int(text))
                return 0

    def detect_and_crop_stickers(self):
        results = self.face_detector(self.image_path)
        boxes = results[0].boxes
        self.face_count = len(boxes)

        for box in boxes:
            top_left_x = int(box.xyxy.tolist()[0][0])
            top_left_y = int(box.xyxy.tolist()[0][1])
            bottom_right_x = int(box.xyxy.tolist()[0][2])
            bottom_right_y = int(box.xyxy.tolist()[0][3])

            width = bottom_right_x - top_left_x
            height = bottom_right_y - top_left_y

            sticker_width = int(width * 1.5)
            sticker_height = height
            sticker_top_left_x = bottom_right_x
            sticker_top_left_y = top_left_y
            sticker_bottom_right_x = sticker_top_left_x + sticker_width
            sticker_bottom_right_y = bottom_right_y

            sticker_region = self.image[sticker_top_left_y:sticker_bottom_right_y, sticker_top_left_x:sticker_bottom_right_x]

            gray_sticker = cv2.cvtColor(sticker_region, cv2.COLOR_BGR2GRAY) 
            _, binary = cv2.threshold(gray_sticker, 30, 255, cv2.THRESH_BINARY)
            contours, _ = cv2.findContours(binary, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
            rectangular_contours = [contour for contour in contours if len(contour) >= 4]

            tmp_sticker_region = sticker_region.copy()
            for contour in rectangular_contours:
                x, y, w, h = cv2.boundingRect(contour)
                cv2.rectangle(tmp_sticker_region, (x, y), (x + w, y + h), (0, 255, 0), 2)
            
            self.find_and_display_numbers(tmp_sticker_region)
            cv2.rectangle(self.image, (int(top_left_x), int(top_left_y)), (int(bottom_right_x), int(bottom_right_y)), (50, 200, 129), 2)
        
        cv2.imwrite(f'{self.folder_path}/Result.jpg', self.image)

    def process(self):
        self.detect_and_crop_stickers()
        self.storage = list(self.storage)
        #write present numbers to file
        file = open(f'{self.folder_path}/presenty.txt','w')
        for i in sorted(self.storage):
            file.write(str(i)+' ')
        file.close()

        db = DBMS(self.storage, self.max_roll, self.folder_path, self.data_sheet, self.data_file)
        db.process()
        return (self.face_count, len(self.storage))
